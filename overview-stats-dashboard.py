# /// script
# requires-python = ">=3.13"
# dependencies = [
#     "altair==6.0.0",
#     "duckdb==1.4.3",
#     "marimo>=0.19.0",
#     "mcp>=1",
#     "openai==2.15.0",
#     "openpyxl==3.1.5",
#     "pandas==2.3.3",
#     "polars[pyarrow]==1.37.1",
#     "pyarrow==22.0.0",
#     "pydantic>=2",
#     "pydantic-ai==1.43.0",
#     "pytest==9.0.2",
#     "pyzmq>=27.1.0",
#     "ruff==0.14.14",
#     "sqlglot==28.6.0",
#     "vegafusion==2.0.3",
#     "vl-convert-python==1.9.0.post1",
# ]
# ///

import marimo

__generated_with = "0.20.2"
app = marimo.App(width="full", app_title="Dutch CRIS / Repositories Dashboard")

async with app.setup(hide_code=True):
    # Initialization code that runs before all other cells
    import marimo as mo
    import micropip

    # Install the packages when running in WASM
    await micropip.install(["polars", "openpyxl"])

    import duckdb
    # engine = duckdb.connect("./data/ducklake.duckdb", read_only=True)
    import altair as alt
    import polars as pl
    import pandas as pd
    import openpyxl


@app.cell(hide_code=True)
def _():
    mo.md("""
    <div style="
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding-bottom: 0.5rem;
        border-bottom: 1px solid #e5e5e5;
        margin-bottom: 1rem;
    ">
        <div>
            <h1 style="margin: 0;">
                Dutch CRIS and Repositories Dashboard
            </h1>
            <div style="color: #666; font-size: 0.9rem;">
                Overview of organisations, data sources, and OpenAIRE compatibility
            </div>
        </div>
        <img
            src="https://www.surf.nl/themes/surf/logo.svg"
            alt="SURF logo"
            style="height: 48px;"
        />
    </div>
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    This dashboard is part of the [**PID to Portal project**](https://communities.surf.nl/en/open-research-information/article/from-pid-to-portal-strengthening-the-open-research-information) from SURF and UNL. The aim is to have all Dutch Research Organisations have their data sources represented correctly in the [Netherlands Research Portal](https://netherlands.openaire.eu/). To claim your Repository / CRIS in the OpenAIRE graph visit [provide.openaire.eu](https://provide.openaire.eu)
    """)
    return


@app.cell(hide_code=True)
def _():
    # Get the Curated Baseline table of Research Organisations in NL

    nl_orgs_baseline = pd.read_excel("https://docs.google.com/spreadsheets/d/e/2PACX-1vTDQiWDIaI1SZkPTMNCovicBhA-nQND1drXoUKvrG1O_Ga3hLDRvmQZao_TvNgmNQ/pub?output=xlsx")
    return (nl_orgs_baseline,)


@app.cell(hide_code=True)
def _():
    # Get the table containing ROR's and OpenAIRE ORG ID's

    orgs_ids_matching = pd.read_excel("https://docs.google.com/spreadsheets/d/e/2PACX-1vTSaXarmKB4RWMlpEDueeMBnwp4_BYJDUwTgBvhqCQ_-hpco9-fa7yZrAIr0T-TIA/pub?output=xlsx")
    return (orgs_ids_matching,)


@app.cell(hide_code=True)
def _(orgs_ids_matching):
    # Add a column with the URL to the Organisation, pointing to the NL research portal

    orgs_ids_matching_with_links = orgs_ids_matching.assign(
        OpenAIRE_ORG_LINK="https://netherlands.openaire.eu/search/organization?organizationId=" + orgs_ids_matching["OpenAIRE_ORG_ID"]
    )
    return (orgs_ids_matching_with_links,)


@app.cell(hide_code=True)
def _(nl_orgs_baseline, orgs_ids_matching_with_links):
    # Merge the baseline table containing RORs, with the table containing ROR's and OpenAIRE ORG ID's

    # SQL join later on was not working in WASM , because the this line produced a polars data frame:
    # SELECT b.full_name_in_English AS name, b.acronym_EN, b.main_grouping AS grouping, m.OpenAIRE_ORG_LINK, m.OpenAIRE_ORG_ID, b.ROR_LINK FROM nl_orgs_baseline b JOIN orgs_ids_matching_with_links m ON b.ROR = m.ROR
    # we replaced it with a merge with Panda's, to keep it Panda's, in order to merge with the data sources data frame later on.

    # keep only the columns you truly need from each side to avoid _x/_y
    left = nl_orgs_baseline[["full_name_in_English", "acronym_EN", "main_grouping", "ROR", "ROR_LINK"]]
    right = orgs_ids_matching_with_links[["ROR", "OpenAIRE_ORG_ID", "OpenAIRE_ORG_LINK"]]

    tmp = left.merge(right, on="ROR", how="inner")

    organisations = tmp.rename(columns={
        "full_name_in_English": "name",
        "main_grouping": "grouping",
    })[["name", "acronym_EN", "grouping", "OpenAIRE_ORG_LINK", "OpenAIRE_ORG_ID", "ROR_LINK"]]
    return (organisations,)


@app.cell(hide_code=True)
def _():
    # Get the DataSources table

    datasources_baseline = pd.read_excel("https://docs.google.com/spreadsheets/d/e/2PACX-1vQwM24DIUWmqbjxaAy62w9w8gNpOMSg5sxmFro-OexCeMzIlyUJh5iVVsVxyrcLkQ/pub?output=xlsx")
    return (datasources_baseline,)


@app.cell(hide_code=True)
def _():
    # Get the OAI-endpoint metrics from the Excel file and load it into a dataframe
    metrics_url = "https://raw.githubusercontent.com/surf-ori/dutch-sources/main/data/nl_orgs_openaire_datasources_with_endpoint_metrics.xlsx"
    datasources_oai_metrics = pd.read_excel(metrics_url)
    return (datasources_oai_metrics,)


@app.cell(hide_code=True)
def _(datasources_baseline):
    # Add a column with the URL to the data source, pointing to the NL research portal

    datasources_url = datasources_baseline.assign(
        OpenAIRE_DataSource_LINK="https://netherlands.openaire.eu/search/dataprovider?datasourceId=" + datasources_baseline["OpenAIRE_DataSource_ID"]
    )
    return (datasources_url,)


@app.cell(hide_code=True)
def _(datasources_baseline, datasources_oai_metrics, datasources_url):
    # Merge datasources_baseline with datasources_oai_metrics and datasources_url on OpenAIRE_DataSource_ID
    # Only add columns from the right tables that are not already present in datasources_baseline

    # Identify columns to add from datasources_oai_metrics
    oai_metrics_cols_to_add = [
        col for col in datasources_oai_metrics.columns
        if col not in datasources_baseline.columns and col != "OpenAIRE_DataSource_ID"
    ]

    # Identify columns to add from datasources_url
    url_cols_to_add = [
        col for col in datasources_url.columns
        if col not in datasources_baseline.columns and col != "OpenAIRE_DataSource_ID"
    ]

    # Merge with datasources_oai_metrics
    datasources = datasources_baseline.merge(
        datasources_oai_metrics[["OpenAIRE_DataSource_ID"] + oai_metrics_cols_to_add],
        on="OpenAIRE_DataSource_ID",
        how="left"
    )

    # Merge with datasources_url
    datasources = datasources.merge(
        datasources_url[["OpenAIRE_DataSource_ID"] + url_cols_to_add],
        on="OpenAIRE_DataSource_ID",
        how="left"
    )

    # Keep only rows with unique OpenAIRE_DataSource_ID
    datasources = datasources.drop_duplicates(subset=["OpenAIRE_DataSource_ID"])
    return (datasources,)


@app.cell(hide_code=True)
def _(datasources, organisations):
    orgs_ds = mo.sql(
        f"""
        SELECT * FROM organisations o FULL JOIN datasources d ON o.openaire_org_id = d.openaire_org_id
        """,
        output=False
    )
    return (orgs_ds,)


@app.cell(hide_code=True)
def _(orgs_ds):
    # Step 1: Get sorted unique values from the specified columns
    # First, we need to get the unique values from the specified columns in the orgs_ds table. We can do this using the unique method provided by polars.

    unique_grouping = orgs_ds["grouping"].unique().sort()
    unique_names = orgs_ds["name"].unique().sort()
    unique_is_geregistreerd = orgs_ds["is_geregistreerd"].unique().sort()
    unique_in_portal = orgs_ds["in portal"].unique().sort()
    unique_wenselijk = orgs_ds["Wenselijk"].unique().sort()
    unique_akkoord_centraal_nl_beheer = orgs_ds["akkoord centraal NL beheer"].unique().sort()
    unique_type = orgs_ds["Type"].unique().sort()
    unique_openaire_compatibility = orgs_ds["openaireCompatibility"].unique().sort()
    unique_oai_status = orgs_ds["oai_status"].unique().sort()


    # Step 2: Create dropdown widgets
    ## Next, we can create dropdown widgets using mo.ui.dropdown. We'll pass the unique values as options to the dropdown.

    grouping_dropdown = mo.ui.dropdown(
        options=["None"] + unique_grouping.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:folder')} Group"
    )

    name_dropdown = mo.ui.dropdown(
        options=["None"] + unique_names.to_list(),
        value="None", # default value 
        label=f"{mo.icon('lucide:landmark')} Org" )

    type_dropdown = mo.ui.dropdown(
        options=["None"] + unique_type.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:type')} Type"
    )

    openaire_compatibility_dropdown = mo.ui.dropdown(
        options=["None"] + unique_openaire_compatibility.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:puzzle')} Compatibility"
    )

    is_geregistreerd_dropdown = mo.ui.dropdown(
        options=["None"] + unique_is_geregistreerd.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:check-square')} Claimed/Registered"
    )

    in_portal_dropdown = mo.ui.dropdown(
        options=["None"] + unique_in_portal.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:globe')} Active In Portal"
    )

    wenselijk_dropdown = mo.ui.dropdown(
        options=["None"] + unique_wenselijk.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:heart')} Required In Portal"
    )

    akkoord_centraal_nl_beheer_dropdown = mo.ui.dropdown(
        options=["None"] + unique_akkoord_centraal_nl_beheer.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:shield')} Managed by SURF"
    )

    oai_status_dropdown = mo.ui.dropdown(
        options=["None"] + unique_oai_status.to_list(),
        value="None",  # default value
        label=f"{mo.icon('lucide:alert-triangle')} OAI Status"
    )

    # Step 3: Create multi-select widget for OAI Endpoint support
    ## We create the multiselect equivalent for the dropdown widgets.

    # Step 3: Multiselect equivalents for ALL dropdown widgets

    grouping_multiselect = mo.ui.multiselect(
        options=unique_grouping.to_list(),
        value=[],
        label=f"{mo.icon('lucide:folder')} Group",
    )

    name_multiselect = mo.ui.multiselect(
        options=unique_names.to_list(),
        value=[],
        label=f"{mo.icon('lucide:landmark')} Organisation",
    )

    type_multiselect = mo.ui.multiselect(
        options=unique_type.to_list(),
        value=[],
        label=f"{mo.icon('lucide:type')} Type",
    )

    openaire_compatibility_multiselect = mo.ui.multiselect(
        options=unique_openaire_compatibility.to_list(),
        value=[],
        label=f"{mo.icon('lucide:puzzle')} Compatibility",
    )

    is_geregistreerd_multiselect = mo.ui.multiselect(
        options=unique_is_geregistreerd.to_list(),
        value=[],
        label=f"{mo.icon('lucide:check-square')} Claimed/Registered",
    )

    in_portal_multiselect = mo.ui.multiselect(
        options=unique_in_portal.to_list(),
        value=[],
        label=f"{mo.icon('lucide:globe')} Active In Portal",
    )

    wenselijk_multiselect = mo.ui.multiselect(
        options=unique_wenselijk.to_list(),
        value=[],
        label=f"{mo.icon('lucide:heart')} Required In Portal",
    )

    akkoord_centraal_nl_beheer_multiselect = mo.ui.multiselect(
        options=unique_akkoord_centraal_nl_beheer.to_list(),
        value=[],
        label=f"{mo.icon('lucide:shield')} Managed by SURF",
    )

    oai_status_multiselect = mo.ui.multiselect(
        options=unique_oai_status.to_list(),
        value=[],
        label=f"{mo.icon('lucide:alert-triangle')} OAI Status",
    )

    ## Create a multi-select widget for OAI Endpoint support

    metadata_support_options = [
        ("detected_support_nl_didl", "NL-DIDL"),
        ("detected_support_oai_dc", "OAI-DC"),
        ("detected_support_oai_openaire", "OAI-OpenAIRE"),
        ("detected_support_oai_cerif_openaire", "OAI-CERIF-OpenAIRE"),
        ("detected_support_openaire_data", "OpenAIRE-Data"),
        ("detected_support_rioxx", "RIOXX")
    ]

    metadata_support_multiselect = mo.ui.multiselect(
        options=[label for _, label in metadata_support_options],
        value=[],  # default value
        label=f"{mo.icon('lucide:link')} Metadata Supported"
    )

    # Step 4: Finally, we can display the dropdown widgets.

    ## Define the filters widget (but don't render it yet)
    filters_widget = mo.vstack(
        [
            mo.md("### Filters"),

            mo.md("---"),
            mo.md("#### Organisations:"),
            grouping_multiselect,
            name_multiselect,

            mo.md("---"),
            mo.md("#### Data Sources:"),
            type_multiselect,
            openaire_compatibility_multiselect,
            is_geregistreerd_multiselect,
            in_portal_multiselect,
            wenselijk_multiselect,
            akkoord_centraal_nl_beheer_multiselect,

            mo.md("---"),
            mo.md("#### OAI Endpoint:"),
            oai_status_multiselect,
            metadata_support_multiselect,
        ],
        gap=1,
    )


    filters_card = mo.vstack(
        [
            mo.md(
                """
                <div style="
                    background: #f5f5f5;
                    padding: 12px;
                    border-radius: 10px;
                    border: 1px solid #e5e5e5;
                ">
                """
            ),
            filters_widget,
            mo.md("</div>"),
        ],
        gap=0,
    )

    ## Make the Sidebar footer
    sidebar_footer = mo.vstack(
        [
            # mo.md(f"**Records in selection:** {num_records}"),
            mo.md(
                """
                <div style="margin-bottom: 8px;">
                    <img
                        src="https://www.surf.nl/themes/surf/logo.svg"
                        alt="SURF logo"
                        style="height: 50px;"
                    />
                </div>
                """
            ),
        ],
        gap=1,
    )


    ## Put the filters card and footer in the sidebar
    mo.sidebar(
        item=filters_card,
        footer=mo.md("[SURF | Open Research Information](https://github.com/surf-ori/dutch-sources)"),
        width="400px",             # optional
    )


    # If you want to use the selected values for further processing, you can access them using the value attribute of the dropdown widgets, like this:

    # selected_grouping = grouping_dropdown.value
    # selected_name = name_dropdown.value
    # selected_is_geregistreerd = is_geregistreerd_dropdown.value
    # selected_in_portal = in_portal_dropdown.value
    # selected_wenselijk = wenselijk_dropdown.value
    # selected_akkoord_centraal_nl_beheer = akkoord_centraal_nl_beheer_dropdown.value
    # selected_openaire_compatibility = openaire_compatibility_dropdown.value
    # selected_oai_status = oai_status_dropdown.value
    # selected_metadata_support = metadata_support_multiselect.value

    # To filter the selected columns on the boolean value True, you can use the following approach:
    # filtered_orgs_ds = orgs_ds.filter(
    #     pl.fold(True, lambda acc, col: acc & pl.col(col), metadata_support_multiselect.value)
    # )
    return (
        akkoord_centraal_nl_beheer_dropdown,
        akkoord_centraal_nl_beheer_multiselect,
        grouping_dropdown,
        grouping_multiselect,
        in_portal_dropdown,
        in_portal_multiselect,
        is_geregistreerd_dropdown,
        is_geregistreerd_multiselect,
        name_dropdown,
        name_multiselect,
        oai_status_dropdown,
        oai_status_multiselect,
        openaire_compatibility_dropdown,
        openaire_compatibility_multiselect,
        type_dropdown,
        type_multiselect,
        wenselijk_dropdown,
        wenselijk_multiselect,
    )


@app.cell(hide_code=True)
def _(
    akkoord_centraal_nl_beheer_dropdown,
    akkoord_centraal_nl_beheer_multiselect,
    grouping_dropdown,
    grouping_multiselect,
    in_portal_dropdown,
    in_portal_multiselect,
    is_geregistreerd_dropdown,
    is_geregistreerd_multiselect,
    name_dropdown,
    name_multiselect,
    oai_status_dropdown,
    oai_status_multiselect,
    openaire_compatibility_dropdown,
    openaire_compatibility_multiselect,
    orgs_ds,
    type_dropdown,
    type_multiselect,
    wenselijk_dropdown,
    wenselijk_multiselect,
):
    # filter the data using the selected values from the dropdown widgets and multiselect widgets

    filtered_orgs_ds = orgs_ds

    if grouping_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("grouping") == grouping_dropdown.value
        )

    if name_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("name") == name_dropdown.value
        )

    if type_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("Type") == type_dropdown.value
        )

    if is_geregistreerd_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("is_geregistreerd") == is_geregistreerd_dropdown.value
        )

    if in_portal_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("in portal") == in_portal_dropdown.value
        )

    if wenselijk_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("Wenselijk") == wenselijk_dropdown.value
        )

    if akkoord_centraal_nl_beheer_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("akkoord centraal NL beheer") == akkoord_centraal_nl_beheer_dropdown.value
        )

    if openaire_compatibility_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("openaireCompatibility") == openaire_compatibility_dropdown.value
        )

    if oai_status_dropdown.value not in (None, "None"):
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("oai_status") == oai_status_dropdown.value
        )

    # WITH MULTI-SELECT FILTERS

    # --- Multi-select filters (apply only when list not empty) ---

    if grouping_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("grouping").is_in(grouping_multiselect.value)
        )

    if name_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("name").is_in(name_multiselect.value)
        )

    if type_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("Type").is_in(type_multiselect.value)
        )

    if is_geregistreerd_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("is_geregistreerd").is_in(is_geregistreerd_multiselect.value)
        )

    if in_portal_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("in portal").is_in(in_portal_multiselect.value)
        )

    if wenselijk_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("Wenselijk").is_in(wenselijk_multiselect.value)
        )

    if akkoord_centraal_nl_beheer_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("akkoord centraal NL beheer").is_in(akkoord_centraal_nl_beheer_multiselect.value)
        )

    if openaire_compatibility_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("openaireCompatibility").is_in(openaire_compatibility_multiselect.value)
        )

    if oai_status_multiselect.value:
        filtered_orgs_ds = filtered_orgs_ds.filter(
            pl.col("oai_status").is_in(oai_status_multiselect.value)
        )

    #   FIX THIS FILTER
    # if metadata_support_multiselect.value:
    #    # Map the selected labels to the actual column names
    #    column_mapping = {label: col for col, label in metadata_support_options}
    #    selected_columns = [column_mapping[label] for label in metadata_support_multiselect.value]
    #    
    #    # Ensure the selected columns are treated as boolean expressions
    #    filtered_orgs_ds = filtered_orgs_ds.filter(
    #        pl.fold(True, lambda acc, col: acc & pl.col(col), selected_columns).all()
    #    )

    num_records = filtered_orgs_ds.height
    return (filtered_orgs_ds,)


@app.cell(hide_code=True)
def _(filtered_orgs_ds):
    mo.stop(filtered_orgs_ds is None)

    # Calculate the required statistics
    total_records = filtered_orgs_ds.height
    ja_is_geregistreerd = filtered_orgs_ds.filter(pl.col("is_geregistreerd") == "Ja").height
    ja_in_portal = filtered_orgs_ds.filter(pl.col("in portal") == "Ja").height
    ja_wenselijk = filtered_orgs_ds.filter(pl.col("Wenselijk") == "Ja").height
    count_openaire_compatible = filtered_orgs_ds.filter(
        pl.col("openaireCompatibility").str.contains("OpenAIRE|compatible", literal=False)
    ).height


    # Create a dictionary to hold the stats
    stats = {
        "# Data Sources": total_records,
        "# Claimed by Research Organisation": ja_is_geregistreerd,
        "# Added to NL Research Portal": ja_in_portal,
        "# Required in NL Research Portal": ja_wenselijk,
        "# OpenAIRE Compatible": count_openaire_compatible,
    }

    # Create the cards
    _cards = [
        mo.stat(
            label=label.replace("_", " "),
            value=value,
            bordered=True,
        )
        for label, value in stats.items()
    ]

    # Title for the section
    _title = "## **Data Source Statistics**"

    # Display the cards
    mo.vstack(
        [
            mo.md(_title),
            mo.hstack(_cards, widths="equal", align="center"),
        ]
    )
    return


@app.cell(hide_code=True)
def _(filtered_orgs_ds):
    mo.stop(filtered_orgs_ds is None)

    # --- OAI endpoint statistics (use unique variable names to avoid marimo redefinitions) ---
    oai_count_status_ok = filtered_orgs_ds.filter(pl.col("oai_status") == "ok").height
    oai_count_status_error = filtered_orgs_ds.filter(pl.col("oai_status") == "error").height

    oai_count_support_nl_didl = filtered_orgs_ds.filter(pl.col("detected_support_nl_didl") == True).height
    oai_count_support_oai_dc = filtered_orgs_ds.filter(pl.col("detected_support_oai_dc") == True).height
    oai_count_support_oai_openaire = filtered_orgs_ds.filter(pl.col("detected_support_oai_openaire") == True).height
    oai_count_support_oai_cerif_openaire = filtered_orgs_ds.filter(pl.col("detected_support_oai_cerif_openaire") == True).height
    oai_count_support_openaire_data = filtered_orgs_ds.filter(pl.col("detected_support_openaire_data") == True).height

    # admin email not blank (handles nulls + whitespace)
    oai_count_admin_email_present = filtered_orgs_ds.filter(
        pl.col("admin email")
          .fill_null("")
          .str.strip_chars()
          .ne("")
    ).height

    oai_stats = {
        "# OAI status = ok": oai_count_status_ok,
        "# OAI status = error": oai_count_status_error,
        "# Supports NL-DIDL": oai_count_support_nl_didl,
        "# Supports OAI-DC": oai_count_support_oai_dc,
        "# Supports OAI-OpenAIRE": oai_count_support_oai_openaire,
        "# Supports OAI-CERIF-OpenAIRE": oai_count_support_oai_cerif_openaire,
        "# Supports OpenAIRE-Data": oai_count_support_openaire_data,
        "# Admin email present": oai_count_admin_email_present,
    }

    oai_cards = [
        mo.stat(
            label=label,
            value=value,
            bordered=True,
        )
        for label, value in oai_stats.items()
    ]

    # build the cards layout (use 2 rows so it looks nicer with many cards)
    oai_cards_layout = mo.vstack(
        [
            mo.hstack(oai_cards[:4], widths="equal", align="center"),
            mo.hstack(oai_cards[4:], widths="equal", align="center"),
        ],
        gap=1,
    )

    mo.accordion(
        {
            "OAI Endpoint Statistics": oai_cards_layout,
        },
        multiple=False,
        lazy=False,  # set True if computing stats is expensive
    )
    return


@app.cell
def _():
    mo.md(r"""
    ## Charts
    """)
    return


@app.cell(hide_code=True)
def _():
    mo.md(r"""
    ### Heatmap | Data Source Type vs Compatibility type

    This map shows the number of system types that are not registered yet. Or systems that are harvested in the the wrong compatibility type. Hover over the cells to see more data.
    """)
    return


@app.cell(hide_code=True)
def _(filtered_orgs_ds):
    # 1) Pivot for the heatmap counts (updated Polars args)
    heatmap_data = (
        filtered_orgs_ds
        .pivot(
            index="Type",
            on="openaireCompatibility",
            values="name",
            aggregate_function="len",
        )
        .fill_null(0)
    )

    # 2) Tooltip counts per Type (from original data)
    tooltip_counts = (
        filtered_orgs_ds
        .group_by("Type")
        .agg(
            (pl.col("is_geregistreerd") == "Ja").sum().alias("is_geregistreerd_count"),
            (pl.col("in portal") == "Ja").sum().alias("in_portal_count"),
            (pl.col("Wenselijk") == "Ja").sum().alias("wenselijk_count"),
        )
    )

    # 3) Join tooltip columns onto the pivoted heatmap table
    heatmap_data = heatmap_data.join(tooltip_counts, on="Type", how="left").fill_null(0)

    # 4) Unpivot (replacement for melt)
    meta_cols = ["Type", "is_geregistreerd_count", "in_portal_count", "wenselijk_count"]
    compat_cols = [c for c in heatmap_data.columns if c not in meta_cols]

    heatmap_long = heatmap_data.unpivot(
        index=meta_cols,
        on=compat_cols,
        variable_name="openaireCompatibility",
        value_name="Count",
    )

    # Convert to pandas for Altair (common in marimo)
    heatmap_long_pd = heatmap_long.to_pandas()

    # ---- Sorting helpers (pandas-side, simple + reliable) ----
    # Sort Y by total count per Type
    heatmap_long_pd["type_total"] = heatmap_long_pd.groupby("Type")["Count"].transform("sum")

    # 5) Plot
    heatmap_chart = (
        alt.Chart(heatmap_long_pd)
        .mark_rect()
        .encode(
            x=alt.X("openaireCompatibility:N", title="OpenAIRE Compatibility"),
            y=alt.Y(
                "Type:N",
                title="Type",
                sort=alt.SortField(field="type_total", order="descending"),
            ),
            color=alt.condition(
                alt.datum.Count > 0,
                alt.Color("Count:Q", title="Count", scale=alt.Scale(scheme="viridis", domainMin=1)),
                alt.value("white"),
            ),
            tooltip=[
                alt.Tooltip("Type:N", title="Type"),
                alt.Tooltip("openaireCompatibility:N", title="OpenAIRE Compatibility"),
                alt.Tooltip("Count:Q", title="Cell Count"),
                alt.Tooltip("type_total:Q", title="Row Total"),
                alt.Tooltip("is_geregistreerd_count:Q", title="is_geregistreerd=Ja"),
                alt.Tooltip("in_portal_count:Q", title="In Portal=Ja"),
                alt.Tooltip("wenselijk_count:Q", title="Wenselijk=Ja"),
            ],
        )
        .properties(
        height=400,
        width="container",
    )
        .configure_axis(labelFontSize=12)
        .configure_legend(titleFontSize=14)
    )

    heatmap_chart
    return


@app.cell(hide_code=True)
def _(filtered_orgs_ds):
    # replace _df with your data source
    type_donut_chart = (
        alt.Chart(filtered_orgs_ds)
        .mark_arc(innerRadius=70)
        .encode(
            color=alt.Color(field='Type', type='nominal'),
            theta=alt.Theta(aggregate='count', type='quantitative'),
            tooltip=[
                alt.Tooltip(aggregate='count'),
                alt.Tooltip(aggregate='count'),
                alt.Tooltip(field='Type')
            ]
        )
        .properties(
            height=290,
            width='container',
            config={
                'axis': {
                    'grid': False
                }
            }
        )
    )
    # type_donut_chart
    return (type_donut_chart,)


@app.cell(hide_code=True)
def _(filtered_orgs_ds):
    # group_donut_chart
    group_donut_chart = (
        alt.Chart(filtered_orgs_ds)
        .mark_arc(innerRadius=70)
        .encode(
            color=alt.Color(field='grouping', type='nominal'),
            theta=alt.Theta(aggregate='count', type='quantitative'),
            tooltip=[
                alt.Tooltip(aggregate='count'),
                alt.Tooltip(aggregate='count'),
                alt.Tooltip(field='grouping')
            ]
        )
        .properties(
            height=290,
            width='container',
            config={
                'axis': {
                    'grid': False
                }
            }
        )
    )
    # group_donut_chart
    return (group_donut_chart,)


@app.cell(hide_code=True)
def _(group_donut_chart, type_donut_chart):
    charts_accordion = mo.accordion(
        {
            "Organisation Grouping": mo.vstack(
                [
                    mo.md("### Organisation Grouping"),
                    mo.md(
                        "This donut chart shows the distribution of datasources of organisations "
                        "by their main grouping."
                    ),
                    group_donut_chart,
                ],
                gap=1,
            ),
            "Type Distribution": mo.vstack(
                [
                    mo.md("### Type Distribution"),
                    mo.md(
                        "This donut chart shows the distribution of datasources by their type."
                    ),
                    type_donut_chart,
                ],
                gap=1,
            ),
        },
        multiple=False,
        lazy=False,
    )
    charts_accordion
    return


@app.cell
def _():
    mo.md(r"""
    ## Table
    To see your organisation represented in the OpenAIRE graph, scroll to the right.
    """)
    return


@app.cell(hide_code=True)
def _(filtered_orgs_ds):
    columns_to_hide = [
        "OpenAIRE_ORG_ID_1",
        "Name_1",
        "contactpersoon (uit kvm)",
        "contact persoon email",
    ]

    public_filtered_orgs_ds = filtered_orgs_ds.drop(columns_to_hide)

    public_filtered_orgs_ds
    return


if __name__ == "__main__":
    app.run()
